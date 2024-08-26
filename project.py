
import openpyxl

# 加载 Excel 工作簿
workbook = openpyxl.load_workbook(r'C:\Users\cyn91\Downloads\test1.xlsx')

# 选择工作表
sheet = workbook['工作表1']  # 将 'Sheet1' 替换为你的工作表名称

# 定义一个函数来获取工作表的所有值
def get_values(sheet):
    arr = []  # 第一层列表
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=37):
        a=" "
        if(row[4].value):a=row[4].value
        obj = f"{row[0].value}的座標為({row[1].value},{row[2].value})，他是{row[3].value}的{a}"
        arr.append(obj)
    return arr
# 获取并打印工作表的所有值
data="現在有一間長9單位、寬5單位的教室，現在我在教室的最西南角(0,0)，面朝北方，教室的西北角座標為(0,5)，教室的東南角座標為(9,0)，教室的東北角座標為(9,5)"
values = get_values(sheet)
for value in values:
    data+=","+value
#print(data) 
#串接GEMINI
import google.generativeai as genai
import os

question=input("輸入問題:")

api_key = 'AIzaSyBCcg0skdWwwG-hBucIvDCLHY9FFtzw9-0'
genai.configure(api_key = api_key)

model = genai.GenerativeModel('gemini-pro')
a='1.標點類問題:使用者身下共有29個標點，此類問題是與標點資訊相關，請根據背景資料，並參考當下方向回答。Ex:(1)我現在的位置在哪裡？Ans:您現在在27號標點，桌子在您的前方50公分處，您的左方有一張沙發(2)最近的出口在哪裡?Ans:您的前方有桌子阻擋，請先右轉直走到底，再向左轉直走到底，即可抵達出口2.物品類問題:此類問題提到背景資料中未提到的物品，需開啟辨識功能原地轉身一圈尋找。Ex:(1)最近椅子在哪裡？(2)這裡有飲水機嗎？Ans: 正在幫您尋找最近的飲水機，請開啟辨識功能旋轉一圈3.人物類問題: 所有提到了人物的問題都屬於此類問題，如人名或周圍有多少人，必須開啟辨識功能原地轉身一圈尋找。Ex: (1)人員位置：附近有我認識的人嗎？Ans: 正在幫您尋找附近有沒有認識的人，請開啟辨識功能並請將鏡頭對準臉部高度旋轉一圈(2)社交互動：周圍是否有其他人？Ans: 正在幫您尋找周圍是否有其他人，請開啟辨識功能旋轉一圈4.請求類問題:此類問題是使用者對於中控中心提出的請求，需聯繫中控中心給予使用者支援。Ex:(1)請問可以派人來這裡幫助我嗎?Ans: 正在幫您聯繫中控中心。'

response = model.generate_content("環境資料:"+data+"環境資料到此結束。問題:「"+question+"」問題結束。請根據問題分類:「"+a+"」判斷問題屬於五種問題的哪一種，先回答問題屬於五種問題的哪一種(回答限定:1.標點類問題；2.物品類問題；3.人物類問題；4.請求類問題)，再參考示範回答Ans的形式，根據環境資料回答問題，計算距離請用((x座標差)*(x座標差)+(y座標差)*(y座標差))^0.5的方式計算，請省略過程直接給我簡潔的答案ex:廁所在你的右前方，距離你5公尺")
print(response.text)

#判斷
keyworda = ["標點類問題"]
keywordb = ["物品類問題"]
keywordc = ["人物類問題"]
keywordd = ["請求類問題"]


b= response.text
c=0

for keyword in keyworda:
    if keyword in b:
        print(f"1")
        # 執行某件事
        break
for keyword in keywordb:
    if keyword in b:
        print(f"2")
        # 執行某件事
        break
for keyword in keywordc:
    if keyword in b:
        print(f"3")
        # 執行某件事
        break
for keyword in keywordd:
    if keyword in b:
        print(f"4")
        # 執行某件事
        break
for keyword in keyworde:
    if keyword in b:
        print(f"5")
        # 執行某件事
        break