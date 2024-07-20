class Gemini:
    import openpyxl

    # 加载 Excel 工作簿
    workbook = openpyxl.load_workbook(r'/home/tkuim-sd/wv/yuno/test1.xlsx')

    # 选择工作表
    sheet = workbook['工作表1']  # 将 'Sheet1' 替换为你的工作表名称

    # 定义一个函数来获取工作表的所有值
    def get_values(sheet):
        arr = []  # 第一层列表
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=37):
            a=" "
            if(row[4].value):a=row[4].value
            obj = f"{row[0].value}的座標為({row[1].value},{row[2].value})，他是{row[3].value}的{a}"
            arr.append(obj)
        return arr
    
    # 获取并打印工作表的所有值
    data="現在有一間長9單位、寬5單位的教室，現在我在教室的最西南角(0,0)，面朝北方，教室的西北角座標為(0,5)，教室的東南角座標為(9,0)，教室的東北角座標為(,5)"
    values = get_values(sheet)
    for value in values:
        data+=","+value
    print(data) 

    
    def get_gemini(text):
        #串接GEMINI
        import google.generativeai as genai
        import os
        question=text

        api_key = 'AIzaSyBCcg0skdWwwG-hBucIvDCLHY9FFtzw9-0'
        genai.configure(api_key = api_key)

        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(Gemini.data+question+"，如果遇到計算距離的問題，請利用勾股定理((長*長+寬*寬)開根號)的方式來計算，如果我的路徑經過了一個無法越過的物品的範圍，那就是會撞到")
        print(response.text)

